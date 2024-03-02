from fastapi import FastAPI
from openpyxl import load_workbook

app = FastAPI()
wb = load_workbook('./app/data.xlsx')
ws = wb['Sheet1']
data = [
    [None, ws['C2'].value, ws['D2'].value],
    [ws['B3'].value, None, ws['D3'].value],
     [ws['B4'].value, ws['C4'].value, None]
]

@app.get("/")
async def root():
    return {"message": "Hello World"}

@app.get("/data")
async def get_data():
    return data

if __name__ == '__main__':
    print(data)